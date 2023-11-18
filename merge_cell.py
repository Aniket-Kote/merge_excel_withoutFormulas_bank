from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def merge_cells_in_excel(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    start_col_nums=[11,13,15,19,32,34,36,38,40]
    end_col_num=[12,14,16,30,33,35,37,39,41]
    
    row_num=[2,3]
    
    for r in row_num:
        for c1,c2 in zip(start_col_nums,end_col_num):
            print(r,r,c1,c2)   
            start_cell = f"{get_column_letter(c1)}{r}"
            end_cell = f"{get_column_letter(c2)}{r}"

            merged_cell_range = f"{start_cell}:{end_cell}"
            sheet.merge_cells(merged_cell_range)

    workbook.save(file_path)