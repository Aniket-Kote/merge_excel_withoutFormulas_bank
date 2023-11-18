# from copy_cell import copy_and_paste, final_copy
import openpyxl

import time
from check_if_exists import check_file_exists



from delete__sheet import delete_worksheet
from merge_cell import merge_cells_in_excel

from multiply_by import multiply_column_by_100

def file_processing(source_input_file,output_file):
    # Example usage:

    print('started')
    check_file_exists("./processed/Final_processed.xlsx")
    start_time = time.time()
    # Import the Excel files
    source_file = openpyxl.load_workbook(source_input_file,data_only=True)
    destination_file = openpyxl.Workbook()

    # Get the worksheet to copy
    source_worksheet = source_file.active

    # Create a new worksheet in the destination file
    destination_worksheet = destination_file.create_sheet('Branch_Master_File')

    # Initialize a flag to keep track if '#REF!' is found
    found_ref = False

    # Copy the contents of the source worksheet to the destination worksheet
    for row in list(source_worksheet.iter_rows()):
        if found_ref:
            break  # Exit the outer loop when '#REF!' is found
        for c, cell in enumerate(row):
            value = cell.value
            # or value=='#DIV/0!'
            if value == '#REF!' or value=='#DIV/0!'  or value == '#N/A' or value=='#NAME?' or value=='#NULL!' or value=='#VALUE!' or value=='#SPILL!' or value=='#SPILL!' or value=='#CALC!' or value=='#NUM!' or value=='#### error':
                # print("sheet error at cell ",cell)
                found_ref = True
                break  # Break the inner loop when '#REF!' is found
    arr=[]
    temp_arr=[]   
    if found_ref==False:
        for row in list(source_worksheet.iter_rows()):
            for c, cell in enumerate(row):
                value = cell.value
                # temp_arr.append(str(value))
            # arr.append(temp_arr)
            # temp_arr=[]
                # print(value)
                destination_worksheet.cell(row[0].row, cell.column).value = value

    if found_ref==False:
        destination_file.save(output_file)

    # final_copy(output_file)
    # perform_copy_and_paste_tasks(output_file)
    merge_cells_in_excel(output_file, 'Branch_Master_File')
    













    # Replace these values with your specific file path, sheet name, and starting row number (5 in your case)
    file_path = output_file
    sheet_name = 'Branch_Master_File'
    

    multiply_column_by_100(file_path, sheet_name) 
    # # column=7
    # for column in columns:
    #     multiply_column_by_100(file_path, sheet_name, starting_row,column)

    delete_worksheet(output_file, 'Sheet')


    print("ended")  
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Execution Time: {execution_time} seconds")


